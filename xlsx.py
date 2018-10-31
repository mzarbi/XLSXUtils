import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Color, Alignment, Border, Side

from xlsx_exceptions import EmptyNameException


class XLSX(Workbook):
    def __init__(self):
        Workbook.__init__(self)
        self.name = None

    def setFilename(self, fname):
        self.name = fname
        return self

    def createSheet(self, sheetName):
        return self.create_sheet(title=sheetName)

    def render(self):
        if self.name == None:
            raise EmptyNameException("Empty name exception")
        else:
            self.save(self.name)


class WorkSheetUtils():

    def __init__(self):
        self.index = 1

    def addTitle(self, sheet, title, idx=1, idxH="A"):
        co = mergedCellsCount(title, 40)
        sheet.merge_cells(multiMergeString(idxH, co, 1, 6))
        sheet[idxH + str(idx)] = title

        font = Font(name='Calibri',
                    size=40,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

        alignment = Alignment(horizontal='center', vertical='center')

        sheet[idxH + str(idx)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='87CEEB')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        sheet[idxH + str(idx)].fill = my_fill
        sheet[idxH + str(idx)].alignment = alignment
        self.index += 6
        return co

    def addSubTitle(self, sheet, subtitle):
        co = mergedCellsCount(subtitle, 28)
        sheet.merge_cells(multiMergeString("A", co, self.index, self.index + 2))
        sheet['A' + str(self.index)] = subtitle

        font = Font(name='Calibri',
                    size=28,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        alignment = Alignment(horizontal='center', vertical='center')


        sheet['A' + str(self.index)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='87CEEB')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        sheet['A' + str(self.index)].fill = my_fill
        sheet['A' + str(self.index)].alignment = alignment

        self.index += 3

    def addTitleBar(self, sheet, title, f1, f2):
        st = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
            ",")

        self.addPicture(sheet, f1)
        co = self.addTitle(sheet, title, idx=1, idxH="C")
        self.addPicture(sheet, f2, st[co + 3] + "1")

    def addDate(self):
        date = datetime.datetime.now()
        d = datetime.datetime.strftime(date, '%B %d, %Y')
        co = mergedCellsCount(d, 12)
        sheet.merge_cells(multiMergeString("A", co, self.index, self.index))
        sheet['A' + str(self.index)] = d

        self.index += 1

    def addSpace(self, sheet):
        self.index += 1

    def addH1(self, sheet, h1):
        pass

    def addH2(self, sheet, h2):
        co = mergedCellsCount(h2, 16)
        sheet.merge_cells(multiMergeString("A", co, self.index, self.index + 1))
        sheet['A' + str(self.index)] = h2

        font = Font(name='Calibri',
                    size=16,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(fill_type=None,
                           start_color='FFFFFF37',
                           end_color='FF000560')

        sheet['A' + str(self.index)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='40e0d0')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        sheet['A' + str(self.index)].fill = my_fill
        sheet['A' + str(self.index)].alignment = alignment

        self.index += 2

    def addH2Form(self, sheet, h2, h22):
        co = mergedCellsCount(h2, 16)
        sheet.merge_cells(multiMergeString("A", co, self.index, self.index + 1))
        sheet['A' + str(self.index)] = h2

        font = Font(name='Calibri',
                    size=16,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(fill_type=None,
                           start_color='FFFFFF37',
                           end_color='FF000560')

        sheet['A' + str(self.index)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        # sheet['A' + str(self.index)].fill = my_fill
        sheet['A' + str(self.index)].alignment = alignment

        st = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
            ",")
        co = mergedCellsCount(h22, 16)

        sheet.merge_cells(multiMergeString(st[co + 2], co, self.index, self.index + 1))
        sheet[st[co + 2] + str(self.index)] = h22

        font = Font(name='Calibri',
                    size=16,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(fill_type=None,
                           start_color='FFFFFF37',
                           end_color='FF000560')

        sheet[st[co + 2] + str(self.index)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        # sheet['A' + str(self.index)].fill = my_fill
        sheet[st[co + 2] + str(self.index)].alignment = alignment

        self.index += 2

    def addH3Form(self, sheet, h2, h22, start='A', ):
        co = mergedCellsCount(h2, 14)
        sheet.merge_cells(multiMergeString(start, co - 2, self.index, self.index))
        sheet[start + str(self.index)] = h2

        font = Font(name='Calibri',
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(fill_type=None,
                           start_color='FFFFFF37',
                           end_color='FF000560')

        sheet[start + str(self.index)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        # sheet['A' + str(self.index)].fill = my_fill

        st = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
            ",")
        co = mergedCellsCount(h22, 16)

        sheet.merge_cells(multiMergeString(st[co + 3], co, self.index, self.index))
        sheet[st[co + 3] + str(self.index)] = h22

        font = Font(name='Calibri',
                    size=14,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        alignment = Alignment(horizontal='center', vertical='center')
        fill = PatternFill(fill_type=None,
                           start_color='FFFFFF37',
                           end_color='FF000560')

        sheet[st[co + 3] + str(self.index)].font = font
        my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)
        # sheet['A' + str(self.index)].fill = my_fill

        self.index += 1

    def addH3(self, sheet, h3):
        pass

    def addTable(self, sheet, header, content):
        headerSize = []
        for i in header:
            headerSize.append(mergedCellsCount(i, 14) + 1)

        self.addHeader(sheet, header, headerSize)
        self.addContent(sheet, content, header, headerSize)

    def addHeader(self, sheet, header, headerSize):
        st = "B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
            ",")
        thick_border = Border(left=Side(style='thick'),
                              right=Side(style='thick'),
                              top=Side(style='thick'),
                              bottom=Side(style='thick'))
        for i in xrange(len(header)):
            start = st[0]
            end = st[headerSize[i] - 1]
            st = st[headerSize[i]:]

            sheet[start + str(self.index)].border = thick_border

            sheet.merge_cells(start + str(self.index) + ":" + end + str(self.index + 1))
            sheet[start + str(self.index)] = header[i]
            sheet[start + str(self.index)].alignment = Alignment(horizontal='center', vertical='center')

        self.index += 2

    def addContent(self, sheet, content, header, headerSize):

        thick_border = Border(left=Side(style='thick'),
                              right=Side(style='thick'),
                              top=Side(style='thick'),
                              bottom=Side(style='thick'))
        for j in xrange(len(content)):
            st = "B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
                ",")
            for i in xrange(len(header)):
                start = st[0]
                end = st[headerSize[i] - 1]
                st = st[headerSize[i]:]

                sheet[start + str(self.index)].border = thick_border

                sheet.merge_cells(start + str(self.index) + ":" + end + str(self.index))
                sheet[start + str(self.index)] = content[j][i]
                sheet[start + str(self.index)].alignment = Alignment(horizontal='center', vertical='center')

            self.index += 1

    def addPicture(self, sheet, filename, pos='A1'):
        img = Image(filename)
        img.width = 125
        img.height = 125
        # img.anchor(sheet.cell('A1'))
        sheet.add_image(img, pos)


def mergedCellsCount(st, fontsize):
    d = {}
    d.update({10: 8})
    d.update({11: 6})
    d.update({12: 6})
    d.update({14: 5})
    d.update({16: 4.5})
    d.update({18: 4})
    d.update({24: 3})
    d.update({28: 3.3})
    d.update({40: 2.8})
    d.update({66: 1.4})
    d.update({96: 0.7})

    return int(len(st) / d[fontsize] + 0.999999)


def mergeString(start, c, v):
    st = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
        ",")
    st = st[st.find(start):]
    e = st[c]
    return start + str(v) + ":" + e + str(v)


def multiMergeString(start, c, v1, v2):
    st = "A,B,C,D,E,F,G,H,I,J,K,L,M,N,O,P,Q,R,S,T,U,V,W,X,Y,Z,AA,AB,AC,AD,AE,AF,AG,AH,AI,AJ,AK,AL,AM,AN,AO,AP,AQ,AR,AS,AT,AU,AV,AW,AX,AY,AZ,BA,BB,BC,BD,BE,BF,BG,BH,BI,BJ,BK,BL,BM,BN,BO,BP,BQ,BR,BS,BT,BU,BV,BW,BX,BY,BZ,CA,CB,CC,CD,CE,CF,CG,CH,CI,CJ,CK,CL,CM,CN,CO,CP,CQ,CR,CS,CT,CU,CV,CW,CX,CY,CZ,DA,DB,DC,DD,DE,DF,DG,DH,DI,DJ,DK,DL,DM,DN,DO,DP,DQ,DR,DS,DT,DU,DV,DW,DX,DY,DZ,EA,EB,EC,ED,EE,EF,EG,EH,EI,EJ,EK,EL,EM,EN,EO,EP,EQ,ER,ES,ET,EU,EV,EW,EX,EY,EZ,FA,FB,FC,FD,FE,FF,FG,FH,FI,FJ,FK,FL,FM,FN,FO,FP,FQ,FR,FS,FT,FU,FV,FW,FX,FY,FZ,GA,GB,GC,GD,GE,GF,GG,GH,GI,GJ,GK,GL,GM,GN,GO,GP,GQ,GR,GS,GT,GU,GV,GW,GX,GY,GZ,HA,HB,HC,HD,HE,HF,HG,HH,HI,HJ,HK,HL,HM,HN,HO,HP,HQ,HR,HS,HT,HU,HV,HW,HX,HY,HZ,IA,IB,IC,ID,IE,IF,IG,IH,II,IJ,IK,IL,IM,IN,IO,IP,IQ,IR,IS,IT,IU,IV,IW,IX,IY,IZ,JA,JB,JC,JD,JE,JF,JG,JH,JI,JJ,JK,JL,JM,JN,JO,JP,JQ,JR,JS,JT,JU,JV,JW,JX,JY,JZ,KA,KB,KC,KD,KE,KF,KG,KH,KI,KJ,KK,KL,KM,KN,KO,KP,KQ,KR,KS,KT,KU,KV,KW,KX,KY,KZ,LA,LB,LC,LD,LE,LF,LG,LH,LI,LJ,LK,LL,LM,LN,LO,LP,LQ,LR,LS,LT,LU,LV,LW,LX,LY,LZ,MA,MB,MC,MD,ME,MF,MG,MH,MI,MJ,MK,ML,MM,MN,MO,MP,MQ,MR,MS,MT,MU,MV,MW,MX,MY,MZ,NA,NB,NC,ND,NE,NF,NG,NH,NI,NJ,NK,NL,NM,NN,NO,NP,NQ,NR,NS,NT,NU,NV,NW,NX,NY,NZ,OA,OB,OC,OD,OE,OF,OG,OH,OI,OJ,OK,OL,OM,ON,OO,OP,OQ,OR,OS,OT,OU,OV,OW,OX,OY,OZ,PA,PB,PC,PD,PE,PF,PG,PH,PI,PJ,PK,PL,PM,PN,PO,PP,PQ,PR,PS,PT,PU,PV,PW,PX,PY,PZ,QA,QB,QC,QD,QE,QF,QG,QH,QI,QJ,QK,QL,QM,QN,QO,QP,QQ,QR,QS,QT,QU,QV,QW,QX,QY,QZ,RA,RB,RC,RD,RE,RF,RG,RH,RI,RJ,RK,RL,RM,RN,RO,RP,RQ,RR,RS,RT,RU,RV,RW,RX,RY,RZ,SA,SB,SC,SD,SE,SF,SG,SH,SI,SJ,SK,SL,SM,SN,SO,SP,SQ,SR,SS,ST,SU,SV,SW,SX,SY,SZ,TA,TB,TC,TD,TE,TF,TG,TH,TI,TJ,TK,TL,TM,TN,TO,TP,TQ,TR,TS,TT,TU,TV,TW,TX,TY,TZ,UA,UB,UC,UD,UE,UF,UG,UH,UI,UJ,UK,UL,UM,UN,UO,UP,UQ,UR,US,UT,UU,UV,UW,UX,UY,UZ,VA,VB,VC,VD,VE,VF,VG,VH,VI,VJ,VK,VL,VM,VN,VO,VP,VQ,VR,VS,VT,VU,VV,VW,VX,VY,VZ,WA,WB,WC,WD,WE,WF,WG,WH,WI,WJ,WK,WL,WM,WN,WO,WP,WQ,WR,WS,WT,WU,WV,WW,WX,WY,WZ,XA,XB,XC,XD,XE,XF,XG,XH,XI,XJ,XK,XL,XM,XN,XO,XP,XQ,XR,XS,XT,XU,XV,XW,XX,XY,XZ,YA,YB,YC,YD,YE,YF,YG,YH,YI,YJ,YK,YL,YM,YN,YO,YP,YQ,YR,YS,YT,YU,YV,YW,YX,YY,YZ,ZA,ZB,ZC,ZD,ZE,ZF,ZG,ZH,ZI,ZJ,ZK,ZL,ZM,ZN,ZO,ZP,ZQ,ZR,ZS,ZT,ZU,ZV,ZW,ZX,ZY,ZZ".split(
        ",")
    st = st[st.index(start):]
    e = st[c]
    return start + str(v1) + ":" + e + str(v2)

if __name__ == "__main__":
    utils = WorkSheetUtils()
    xlsx = XLSX().setFilename("file.xlsx")

    # sheet = xlsx.createSheet("sheet1")
    sheet = xlsx.active
    utils.addTitleBar(sheet, "NEST Optics : Strategic Planner", 'img/sotetel.jpeg', 'img/alcatel.png')
    utils.addDate()
    utils.addSpace(sheet)
    utils.addSpace(sheet)

    utils.addSubTitle(sheet, "Scenario Settings")
    utils.addSpace(sheet)
    utils.addH2Form(sheet, "Name            ", "Scenario 1   ")
    utils.addH2Form(sheet, "Project         ", "Project 1    ")
    utils.addH2Form(sheet, "User            ", "med Zied Arbi")
    utils.addH2Form(sheet, "Gouvernorate    ", "Bizerte      ")
    utils.addH2Form(sheet, "Country         ", "Tunisia      ")
    utils.addSpace(sheet)
    utils.addH2(sheet, "National Scale   ")
    utils.addSpace(sheet)
    utils.addH3Form(sheet, "Temporal Horizon    (Years)  ", "Scenario 1   ", start='B')
    utils.addH3Form(sheet, "Budget Per year    (TND)    ", "Scenario 1   ", start='B')
    utils.addH3Form(sheet, "Budget Fluctuation (%)      ", "Scenario 1   ", start='B')
    utils.addH3Form(sheet, "Projects Per Year  (Project)", "Scenario 1   ", start='B')
    utils.addSpace(sheet)
    utils.addH2(sheet, "Regional Scale  ")
    utils.addSpace(sheet)
    utils.addH3Form(sheet, "Temporal Horizon    (Years)  ", "Scenario 1   ", start='B')
    utils.addH3Form(sheet, "Budget Per year    (TND)    ", "Scenario 1   ", start='B')
    utils.addH3Form(sheet, "Budget Fluctuation (%)      ", "Scenario 1   ", start='B')
    utils.addH3Form(sheet, "Projects Per Year  (Project)", "Scenario 1   ", start='B')
    utils.addSpace(sheet)

    utils.addSubTitle(sheet, "Exported Results ")
    utils.addSpace(sheet)
    utils.addH2(sheet, "Results         ")
    utils.addSpace(sheet)
    utils.addTable(sheet,
                   ["Zone", "technology", "Year", "National Cost", "Regional Cost", "Annual Revenu", "Global Revenu",
                    "Gain", "ROI", "Profit after"],
                   [["Zone", "technology", "Year", "National Cost", "Regional Cost", "Annual Revenu", "Global Revenu",
                     "Gain", "ROI", "Profit after"]
                       ,
                    ["Zone", "technology", "Year", "National Cost", "Regional Cost", "Annual Revenu", "Global Revenu",
                     "Gain", "ROI", "Profit after"]
                       ,
                    ["Zone", "technology", "Year", "National Cost", "Regional Cost", "Annual Revenu", "Global Revenu",
                     "Gain", "ROI", "Profit after"]])
    utils.addSpace(sheet)
    utils.addH2(sheet, "Bill Of Material")
    utils.addSpace(sheet)
    utils.addTable(sheet,
                   ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                   [["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                    ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                    ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                    ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                    ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                    ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"],
                    ["Year", "Zone", "Feature", "Reference", "Manufacturer", "Quantity", "Price"]
                    ])
    utils.addSpace(sheet)

    '''
    utils.addTitle(sheet, "hello")
    utils.addSpace(sheet)
    utils.addSubTitle(sheet, "Qos Design")
    utils.addPicture(sheet,'img/sotetel.jpeg')
    '''
    xlsx.render()

